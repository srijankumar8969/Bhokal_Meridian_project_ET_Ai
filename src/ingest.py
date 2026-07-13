"""
Teammate C's job (Task 3, 4, 6): rip text out of every file type, chunk it,
hand a clean structured list of chunks down the pipeline.

Supports: PDF, XLSX, TXT (email), PNG/JPG (OCR for P&ID tags).
Output: data/chunks.json -- a flat list of {id, source_file, text, doc_type}
"""
import os
import json
import glob
from pypdf import PdfReader
from openpyxl import load_workbook
import pytesseract
from PIL import Image

DATA_DIR = "/home/claude/meridian/data"
OUT_PATH = "/home/claude/meridian/data/chunks.json"

CHUNK_SIZE = 500  # characters
CHUNK_OVERLAP = 50


def extract_pdf(path):
    reader = PdfReader(path)
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    return text


def extract_xlsx(path):
    wb = load_workbook(path, data_only=True)
    lines = []
    for ws in wb.worksheets:
        headers = None
        for row in ws.iter_rows(values_only=True):
            if headers is None:
                headers = row
                continue
            if all(v is None for v in row):
                continue
            row_desc = ", ".join(
                f"{h}: {v}" for h, v in zip(headers, row) if v is not None
            )
            lines.append(row_desc)
    return "\n".join(lines)


def extract_txt(path):
    with open(path, "r") as f:
        return f.read()


def extract_image(path):
    img = Image.open(path)
    return pytesseract.image_to_string(img)


def chunk_text(text, size=CHUNK_SIZE, overlap=CHUNK_OVERLAP):
    """Simple sliding-window character chunker. Splits on paragraph boundaries
    where possible so entities don't get sliced in half."""
    paragraphs = [p.strip() for p in text.split("\n") if p.strip()]
    chunks = []
    current = ""
    for para in paragraphs:
        if len(current) + len(para) < size:
            current = (current + " " + para).strip()
        else:
            if current:
                chunks.append(current)
            # if a single paragraph is itself bigger than size, hard-split it
            while len(para) > size:
                chunks.append(para[:size])
                para = para[size - overlap:]
            current = para
    if current:
        chunks.append(current)
    return chunks


def process_file(path):
    ext = os.path.splitext(path)[1].lower()
    fname = os.path.basename(path)
    if ext == ".pdf":
        text, doc_type = extract_pdf(path), "pdf"
    elif ext == ".xlsx":
        text, doc_type = extract_xlsx(path), "excel"
    elif ext == ".txt":
        text, doc_type = extract_txt(path), "email"
    elif ext in (".png", ".jpg", ".jpeg"):
        text, doc_type = extract_image(path), "image_ocr"
    else:
        print(f"Skipping unsupported file: {fname}")
        return []

    raw_chunks = chunk_text(text)
    return [
        {
            "id": f"{fname}::chunk{i}",
            "source_file": fname,
            "doc_type": doc_type,
            "text": c,
        }
        for i, c in enumerate(raw_chunks)
    ]


def main():
    all_chunks = []
    files = sorted(glob.glob(os.path.join(DATA_DIR, "*")))
    files = [f for f in files if not f.endswith(".json")]
    for path in files:
        chunks = process_file(path)
        print(f"{os.path.basename(path)}: {len(chunks)} chunk(s)")
        all_chunks.extend(chunks)

    with open(OUT_PATH, "w") as f:
        json.dump(all_chunks, f, indent=2)
    print(f"\nTotal: {len(all_chunks)} chunks written to {OUT_PATH}")


if __name__ == "__main__":
    main()
